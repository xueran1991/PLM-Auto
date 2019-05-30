from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Color, PatternFill
import os
import glob
import re
import numpy as np
import pandas as pd


class PLM_auto():

    def __init__(self, base_path):
        # 根模板存储路径
        self.base_path = base_path

        #大类：中类
        self.primary_class = {'电气':set()}
        #中类：次中类
        self.medium_class = {}
        #次中类：小类
        self.detailed_class = {}
        #小类编码
        self.class_code = {}
        
        #品牌和翻译的字典
        self.brands = {}
        self.trans = {}
        #材质的字典
        self.materials = {}
        #读入的数据
        self.data = None
        self.dataframe = None
        self.attribs = None
        self.attribs_necessity = None
        #当前任务
        self.job = ''
        self.job_code = ''
        self.mclass = '' # 中类名称
        self.sclass = '' # 次中类名称
        self.write_path = ''

    def get_class(self, file):
        ## 提取分类类别
        Y201_ruls_path = self.base_path + file
        Y201_ruls_wb = load_workbook(Y201_ruls_path)
        Y201_ruls_ws = Y201_ruls_wb['物料分类表 R22']

        # C: 大类名称 电气……
        # E：中类名称 传感器……
        # F: 次中类名称 连续物位……
        # G：小类名称 雷达……        # H：小类代号 A300……
        C, D, E, F, G, H = Y201_ruls_ws['C:H']        

        for c, e, f, g, h  in zip(C, E, F, G, H):
            if c.value == '电气' and g.value != None:
                self.primary_class['电气'].add(e.value)
                # 存储中类：次中类
                try:
                    self.medium_class[e.value].add(f.value)
                except:
                    self.medium_class[e.value] = set()
                    self.medium_class[e.value].add(f.value)
                # 存储 次中类：小类
                try:
                    self.detailed_class[f.value].add(g.value)
                except:
                    self.detailed_class[f.value] = set()
                    self.detailed_class[f.value].add(g.value)
                # 存储 小类名称：小类编码
                self.class_code[g.value] = h.value

    def get_brands(self, file):
        # 提取品牌库信息
        brands_path =self.base_path + file # 最终品牌库的路径
        brands_file = open(brands_path, 'r')
        brs = brands_file.readlines()
        self.brands = {}
        for br in brs:
            ls = re.split("\s", br)
            if ls[0] != '':
                if ls[0] not in self.brands.keys() \
                and ls[1] not in self.brands.values():

                    self.brands[ls[0]] = ls[1]
                else:
                    print('品牌库异常：', ls[0], ls[1], '出现重复项')

    def get_material(self, file):
        # 提取材质库信息
        m_path = self.base_path + file
        m_file = open(m_path, 'r')
        ms = m_file.readlines()
        materials = {}
        for m in ms:
            m = re.split("\s", m)
            if m[0] != '' :
                materials[m[0]] = m[1]

        m_file.close()
        self.materials = materials

                                            
    def create_class_path(self, class_job='传感器'):
        ### 创建小类路径
        for m_class in self.medium_class[class_job]:
            for d_class in self.detailed_class[m_class]:
                m_class = m_class.replace("/", '')
                d_class = d_class.replace("/", '')
                path = class_job + '/' + m_class + '/' + d_class
                
                if not os.path.exists(path):
                    os.makedirs(path)
                    print('已创建路径：', path)      
                    
    def get_translation(self, file):
        ### 提取翻译库数据
        trans_path = file
        trans_file = open(trans_path, 'r')
        trans = trans_file.readlines()
        translation = {}
        for tran in trans:
            tr = re.split("\s", tran)
            if tr[0] != '' :
                translation[tr[0]] = tr[1]

        trans_file.close()
        return translation            

    def refresh_translation(self, file):
        ### 更新翻译库，将需要导入的属性加入翻译库
        current_trans = self.get_translation(file)
        trans_file = open(file, 'a+')

        for attrib in self.attribs:
            if attrib not in current_trans:
                trans_file.write(attrib+'\n')
        trans_file.close()
        
        new_trans = self.get_translation(file)
        self.trans = new_trans.copy()
        # for key in new_trans.keys():
        #     if new_trans[key] == '':
        #         print(key, ':翻译缺失')

    def get_job_class(self):
        ### 获取当前任务的中类名称和大类名称
        for k, v in self.detailed_class.items():
            if self.job in v:
                self.sclass = k
                break
        for k, v in self.medium_class.items():
            if self.sclass in v:
                self.mclass = k
                break

    def get_data(self, file_path):
        ### 读入原始物料表单数据
        if file_path[-5:] == '.xlsx':
            wb = load_workbook(file_path, 'r')
            ws = wb.active
            data =  np.array(list(ws.values))
            wb.close
            self.data = data 
        elif file_path[-4:] == '.txt':
            txt = open(file_path, 'r')
            txt = txt.readlines()
            data = []
            for row in txt:    
                d = row.replace("\n",'')
                data.append(re.split("\t", d))
            self.data = np.array(data)            
        else:
            raise ValueError
            
        # 提取当前任务的名称（同小类名称）
        # self.job = file_path.split("\\")[-1].split(".")[0]

        # 写入地址是输入文件的路径 + 小类文件夹
        # self.write_path = file_path.replace(self.job+'.txt', self.job+'\\')


        self.dataframe = pd.DataFrame(data=self.data[2:, 1:], \
            columns=self.data[1, 1:], index=self.data[2:, 0])
        # 属性
        self.attribs = self.dataframe.columns[2:]
        # 属性是否必填
        self.attribs_necessity = self.data[0][3:]

        # 小类编码
        # 通过dataframe获取分类名称
        self.job = self.dataframe['分类名称'].unique()[0]
        self.job_code = self.class_code[self.job] 
        # 写入地址是输入文件的路径 + 小类文件夹
        self.write_path = os.path.dirname(file_path) + '\\' + self.job_code + "-" + self.job + '\\'
        if not os.path.exists(self.write_path):
            os.makedirs(self.write_path)
            print('已创建路径：', self.write_path) 

    def create_t02(self):
        ### 生成 02-必填属性
        self.get_job_class()

        t02_header = '#===============电气与自动化================\n\n'
        t02_header += ('#--' + self.mclass + '--' + self.sclass + '--' + self.job + '\n')
        t02_content = self.job_code + '=mySupplierModelSpec'

        # 遍历所有属性，添加必要属性
        for attrib, attrib_necessity in zip(self.attribs, self.attribs_necessity):
            # txt读出的数值为str型，excel读出的为int型
            if attrib_necessity in [1, '1']:
                t02_content += ',my' + self.job_code + '_' + self.trans[attrib]

        # 完成02模板
        t02 = t02_header + t02_content


        t02_file = open(self.write_path+'迈安德物料定义-02必填属性-'+self.job+'.txt', 'w')
        t02_file.write(t02)
        t02_file.close()

    def create_t03(self):
        ### 生成03-属性组合

        t03_header = "#====各分类组合属性集===========\n\
#====分类码.组合属性代码+组合顺序=子属性名\n\
#====物料简称:cassShortDescription\n\
#====描述:myDesc\n\n"
        t03_header += ('#--' + self.mclass + '--' + self.sclass + '--' + self.job + '\n\n')

        t03_content1 = '#l表示：PLM物料名称:ERP物料名称\n' + self.job_code \
                        + '.l0={cassShortDescription}\n\n'

        t03_content2 = '#d：ERP物料名称:物料简称\n' + self.job_code \
                        + '.d0={cassShortDescription}\n\n'

        t03_content3 = "#s：ERP型号规格:供应商型号规格\n" + self.job_code \
                        + ".s0={mySupplierModelSpec}\n\n"
        t03 = t03_header + t03_content1 + t03_content2 + t03_content3

        t03_file = open(self.write_path+'迈安德物料定义-03属性组合-'+self.job+'.txt', 'w')
        t03_file.write(t03)
        t03_file.close()

    def create_t01(self):
        ### 生成 01-属性表 

        # 创建Excel文件
        wb = Workbook()
        ws_df = wb.create_sheet("定义属性")
        ws_rg = wb.create_sheet("Range值")

        # 生成定义属性工作表
        ws_df_header = ['模块','中文名称','分类码','英文名字','定义属性','属性类型','翻译脚本','创建属性脚本']

        # 属性数目
        attr_num = len(self.attribs)
        ws_df_cA = ['物料属性'] * attr_num
        ws_df_cB = self.attribs
        ws_df_cC = [self.job_code] * attr_num
        ws_df_cD = [self.trans[x] for x in ws_df_cB]
        ws_df_cE = ['="my"&C{}&"_"&D{}'.format(i+2,i+2) for i in range(attr_num)]
        ws_df_cF = ['string'] * attr_num
        ws_df_cG = ['=CONCATENATE("emxFramework.Attribute.",E{}," = ",B{},)'.format(i+2,i+2) for i in range(attr_num)]
        cH_string = """=CONCATENATE("#"&B{}&"
add attribute "&E{}&"
  type "&F{}&"
  description '' default ''
  property application value MyandeCentral
  property installer value cass
  property 'original name' value "&E{}&"
  property 'installed date' value 05-01-2018
  property version value 1.0;
mod prog eServiceSchemaVariableMapping.tcl add property attribute_"&E{}&" to att "&E{}&";")"""

        ws_df_cH = [cH_string.format(i,i,i,i,i,i) for i in range(2, attr_num+2)]

        df = pd.DataFrame(columns=ws_df_header)
        df['模块'] = ws_df_cA
        df['中文名称'] = ws_df_cB
        df['分类码'] = ws_df_cC
        df['英文名字'] = ws_df_cD
        df['定义属性'] = ws_df_cE
        df['属性类型'] = ws_df_cF
        df['翻译脚本'] = ws_df_cG
        df['创建属性脚本'] = ws_df_cH

        
        #写入行数据
        for r in dataframe_to_rows(df, index=False, header=True):
            ws_df.append(r)

        # 调整列宽
        ws_df.column_dimensions['A'].width = 9
        ws_df.column_dimensions['B'].width = 15
        ws_df.column_dimensions['C'].width = 9
        ws_df.column_dimensions['D'].width = 25
        ws_df.column_dimensions['E'].width = 30
        ws_df.column_dimensions['F'].width = 9
        ws_df.column_dimensions['G'].width = 80
        ws_df.column_dimensions['H'].width = 50

        #调整背景颜色
        patt1 = PatternFill(start_color='FF92D050',
                   fill_type='solid')
        patt2 = PatternFill(start_color='FF00B050',
                   fill_type='solid')

        for i in range(1, len(ws_df_cG)+3):
            ws_df['G'+str(i)].fill = patt1
            ws_df['H'+str(i)].fill = patt2
            
        ##---------------------------------------------------------
        ## 生成Range值工作表
        ws_rg_header = ['固定属性','分类码','英文名字','定义属性','-','值的前台显示','值(不能有中文)','Range值脚本','Range值翻译']
        ws_rg_cC = []
        ws_rg_cF = []
        ws_rg_cG = []
        range_value = []
        
        for attrib, attrib_necessity in zip(self.attribs, self.attribs_necessity):

            # 检查翻译库
            if self.trans[attrib] == '':
                print('+---翻译库缺失--', attrib)

            ## 遍历dataframe，生成C，F, G列 和 属性的值
            i = 1 # 每种普通属性的值从1开始计数
            # txt读出的数值为str型，excel读出的为int型
            if attrib_necessity in [1, '1']:
            ## 如果属性是必填属性，在ws_rg_cC中存储属性的枚举，
            ## 并在每一个属性后添加一个空的属性值
                for value in self.dataframe[attrib].unique():
                    if value != '' and value != None:
                        ws_rg_cC.append(self.trans[attrib])
                        # 在F列写入属性的名称
                        ws_rg_cF.append(value)

                    # 在G列填入品牌库编码
                    if '品牌' in attrib:
                        try:
                            ws_rg_cG.append(self.brands[value])
                        except:
                            ws_rg_cG.append('')
                            print('+---品牌库缺失--', value)

                    # 在G列填入材质库编码 ##---取消材质库关联
                    # elif '材质' in attrib:
                    #     try:
                    #         ws_rg_cG.append(self.materials[value])
                    #     except:
                    #         print('+---材质库缺失--', value)
                    #         ws_rg_cG.append('')

                    # 在G列填入其他属性的值
                    else:
                        ws_rg_cG.append(self.trans[attrib].lower() + str(i))
                        i += 1
            
                # 每种必填属性穷举后添加一行，这行在cC中显示属性的翻译，属性值和前台显示为空
                ws_rg_cC.append(self.trans[attrib])
                ws_rg_cF.append('')
                ws_rg_cG.append('')
        ## 生成其他列，写入表格公式或值
        # 必填属性枚举的数目
        nece_attrib_num = len(ws_rg_cF)
        ws_rg_cA = ['mod attr'] * nece_attrib_num
        ws_rg_cB = [self.job_code] * nece_attrib_num
        ws_rg_cD = ['="my"&B{}&"_"&C{}'.format(i,i) for i in range(2,nece_attrib_num+2)]
        ws_rg_cE = ['add range'] * nece_attrib_num

        # H、I列写入公式
        cH_string = """=CONCATENATE(A{}," ",D{}," ",E{}," ","="," ","'",G{},"'",";")"""
        ws_rg_cH = [cH_string.format(i,i,i,i) for i in range(2, nece_attrib_num+2)]
        cI_string = '=CONCATENATE("emxFramework.Range.",D{},,".",G{}," ="," ",F{})'
        ws_rg_cI = [cI_string.format(i,i,i) for i in range(2, nece_attrib_num+2)]

        # 将数据传入df
        df = pd.DataFrame(columns=ws_rg_header)
        df['固定属性'] = ws_rg_cA
        df['分类码'] = ws_rg_cB
        df['英文名字'] = ws_rg_cC
        df['定义属性'] = ws_rg_cD
        df['-'] = ws_rg_cE
        df['值的前台显示'] = ws_rg_cF
        df['值(不能有中文)'] = ws_rg_cG
        df['Range值脚本'] = ws_rg_cH
        df['Range值翻译'] = ws_rg_cI

        # 将df写入Excel
        for r in dataframe_to_rows(df, index=False, header=True):
            ws_rg.append(r)

        # 调整格式
        ws_rg.column_dimensions['A'].width = 9
        ws_rg.column_dimensions['B'].width = 6
        ws_rg.column_dimensions['C'].width = 28
        ws_rg.column_dimensions['D'].width = 37
        ws_rg.column_dimensions['E'].width = 9
        ws_rg.column_dimensions['F'].width = 40
        ws_rg.column_dimensions['G'].width = 20
        ws_rg.column_dimensions['H'].width = 100
        ws_rg.column_dimensions['I'].width = 150

        for i in range(1, len(ws_rg_cH)+3):
            ws_rg['H'+str(i)].fill = patt1
            ws_rg['I'+str(i)].fill = patt2
            
        wb.active = 1
        wb.save(self.write_path+'迈安德物料定义-01属性表-外购设备-'+self.job+".xlsx")
        wb.close()


    def create_t04(self):
        ### 生成 04-导入模板

        parts_num = int(self.dataframe.index[-1])
        export_header0 = ['', 'name', 'PartFamily:name|attribute[Title]','cassLongDescription', \
                         'description','cassSpecModel','cassShortDescription','myDesc','mySupplierModelSpec']

        export_header1 = ['序号','编码','分类名称','PLM物料名称','ERP物料名称', \
                          'ERP型号规格','物料简称','描述','供应商型号规格']
        # 英文表头和中文表头
        export_header0 += ['my'+self.job_code+'_'+self.trans[at] for at in self.attribs]
        export_header1 += list(self.attribs)

        df_export = pd.DataFrame(columns=export_header1)
        df_export['序号'] = range(1, parts_num+1)
        df_export['分类名称'] = [self.job+'|'+self.job_code] * parts_num
        df_export['ERP物料名称'] = df_export['PLM物料名称'] = df_export['物料简称'] = [self.job]* parts_num
        df_export['供应商型号规格'] = df_export['ERP型号规格'] = list(self.dataframe['供应商型号规格'])[:parts_num]

        for attrib in self.attribs:
            df_export[attrib] = list(self.dataframe[attrib])[:parts_num]

        encodings = [self.job_code + '-' + str(10000+i) + '-{}' for i in range(1, parts_num+1)]
        df_export['编码'] = [encoding.format(self.brands[x]) for encoding, x in zip(encodings, df_export['品牌'])]    
            
        wb = Workbook()
        ws = wb.active
        ws.append(export_header0)

        for r in dataframe_to_rows(df_export, index=False, header=True):
            ws.append(r)
            
        wb.save(self.write_path + '迈安德物料定义-04物料导入模板-' + self.job + ".xlsx")
        wb.close()